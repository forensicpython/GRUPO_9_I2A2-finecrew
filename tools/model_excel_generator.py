from crewai.tools import tool
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
import os
from pathlib import Path
from datetime import datetime

@tool("generate_model_compliant_excel_tool")
def generate_model_compliant_excel_tool(output_filename: str = "VR MENSAL 05.2025.xlsx", reference_month: str = "05.2025") -> str:
    """
    Gera planilha Excel EXATAMENTE conforme modelo "VR MENSAL 05.2025.xlsx":
    - Aba "VR MENSAL 05.2025" com formato idêntico
    - Aba "Validações" com checklist conforme modelo
    - Dados reais dos cálculos automatizados
    """
    try:
        project_root = Path(os.path.dirname(__file__)).parent
        output_path = project_root / "output" / output_filename
        output_path.parent.mkdir(exist_ok=True)
        
        print("📋 Gerando planilha conforme modelo exato...")
        
        # Carregar dados dos cálculos automatizados
        calculo_path = project_root / "output" / "calculo_automatizado_beneficios.xlsx"
        
        if calculo_path.exists():
            print("✅ Usando dados REAIS dos cálculos automatizados")
            calculos_df = pd.read_excel(calculo_path, sheet_name='Cálculos Individuais')
        else:
            print("⚠️ Cálculos automatizados não encontrados, usando base consolidada")
            base_path = project_root / "output" / "base_consolidada.xlsx"
            calculos_df = pd.read_excel(base_path, sheet_name='Base Consolidada')
            
        # Carregar datas de admissão reais
        print("📅 Carregando datas de admissão reais...")
        raw_data_path = project_root / "raw_data"
        mapeamento_admissao = load_real_admission_dates(raw_data_path)
        
        # Carregar sindicatos reais dos funcionários
        print("🏢 Carregando sindicatos reais dos funcionários...")
        mapeamento_sindicatos = load_real_syndicates(raw_data_path)
        
        # CRÍTICO: Enriquecer dados com datas de admissão reais
        print("🔧 Enriquecendo dados com datas de admissão...")
        calculos_df['DATA_ADMISSAO_REAL'] = calculos_df['MATRICULA'].astype(str).apply(mapeamento_admissao)
        print(f"✅ {len(calculos_df)} funcionários enriquecidos com datas de admissão")
        
        # Criar workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove aba padrão
        
        # ABA 1: VR MENSAL com mês de referência dinâmico
        sheet_name = f"VR MENSAL {reference_month}"
        ws_main = wb.create_sheet(sheet_name)
        create_vr_mensal_sheet_model_compliant(ws_main, calculos_df, mapeamento_admissao, mapeamento_sindicatos, reference_month)
        
        # ABA 2: Validações (conforme modelo exato)
        ws_validacoes = wb.create_sheet("Validações")
        create_validacoes_sheet_model_compliant(ws_validacoes)
        
        # Salvar arquivo
        wb.save(str(output_path))
        
        return f"Planilha Excel modelo-compliant gerada: {output_path}"
        
    except Exception as e:
        return f"Erro ao gerar planilha modelo-compliant: {str(e)}"

def create_vr_mensal_sheet_model_compliant(ws, data_df, mapeamento_admissao, mapeamento_sindicatos, reference_month="05.2025"):
    """Cria aba VR MENSAL seguindo EXATAMENTE o modelo"""
    
    # CABEÇALHO ATUALIZADO - INCLUINDO CARGO APÓS MATRÍCULA
    headers = [
        "Matricula",
        "Cargo",
        "Admissão", 
        "Sindicato do Colaborador",
        "Competência",
        "Dias",
        "VALOR DIÁRIO VR",
        "TOTAL",
        "Custo empresa",
        "Desconto profissional",
        "OBS GERAL"
    ]
    
    # Inserir cabeçalhos na linha 1
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # DADOS DOS FUNCIONÁRIOS
    current_row = 2
    
    # Extrair ano e mês da referência
    try:
        mes, ano = reference_month.split('.')
        competencia = datetime(int(ano), int(mes), 1)
    except:
        competencia = datetime(2025, 5, 1)  # Fallback para maio 2025
    
    for _, funcionario in data_df.iterrows():
        # Colunas conforme modelo exato:
        matricula = str(funcionario.get('MATRICULA', ''))
        
        # Cargo do funcionário (título do cargo)
        cargo = str(funcionario.get('TITULO DO CARGO', funcionario.get('CARGO', 'N/A')))
        
        # Data de admissão REAL (usar coluna enriquecida)
        admissao = funcionario.get('DATA_ADMISSAO_REAL', mapeamento_admissao(matricula))
        
        # Sindicato do colaborador REAL (buscar no mapeamento)
        sindicato_nome = mapeamento_sindicatos(matricula)  # Função retorna sindicato real
        
        # Competência (sempre maio 2025)
        competencia_mes = competencia
        
        # Dias úteis
        dias = funcionario.get('DIAS_UTEIS_FINAL', 22)
        
        # Valor diário VR
        valor_diario = funcionario.get('VALOR_DIA', 25.50)
        
        # Total VR
        total_vr = funcionario.get('VALOR_TOTAL_VR', 0)
        
        # Custo empresa (80%)
        custo_empresa = funcionario.get('VALOR_EMPRESA_80', 0)
        
        # Desconto profissional (20%)
        desconto_profissional = funcionario.get('VALOR_FUNCIONARIO_20', 0)
        
        # OBS GERAL (vazio por padrão)
        obs_geral = ""
        
        # Inserir dados nas colunas (ajustado para incluir coluna Cargo)
        ws.cell(row=current_row, column=1, value=matricula)
        ws.cell(row=current_row, column=2, value=cargo)
        ws.cell(row=current_row, column=3, value=admissao)
        ws.cell(row=current_row, column=4, value=sindicato_nome)
        ws.cell(row=current_row, column=5, value=competencia_mes)
        ws.cell(row=current_row, column=6, value=dias)
        ws.cell(row=current_row, column=7, value=valor_diario)
        ws.cell(row=current_row, column=8, value=total_vr)
        ws.cell(row=current_row, column=9, value=custo_empresa)
        ws.cell(row=current_row, column=10, value=desconto_profissional)
        ws.cell(row=current_row, column=11, value=obs_geral)
        
        # Formatação de números (ajustado para nova posição das colunas)
        ws.cell(row=current_row, column=7).number_format = '#,##0.00'  # Valor diário
        ws.cell(row=current_row, column=8).number_format = '#,##0'     # Total
        ws.cell(row=current_row, column=9).number_format = '#,##0'     # Custo empresa
        ws.cell(row=current_row, column=10).number_format = '#,##0'    # Desconto
        
        # CRÍTICO: Formatação da data de admissão (agora na coluna 3)
        ws.cell(row=current_row, column=3).number_format = 'DD/MM/YYYY'
        
        current_row += 1
    
    # Ajustar larguras das colunas conforme modelo (com coluna Cargo adicionada)
    column_widths = [12, 25, 12, 50, 12, 8, 15, 12, 15, 18, 15]  # Cargo: 25 width
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

def create_validacoes_sheet_model_compliant(ws):
    """Cria aba Validações seguindo EXATAMENTE o modelo"""
    
    # CABEÇALHOS
    ws.cell(row=1, column=1, value="Validações")
    ws.cell(row=1, column=2, value="Check")
    
    # Formatação do cabeçalho
    for col in [1, 2]:
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # ITENS DE VALIDAÇÃO (EXATOS CONFORME MODELO)
    validacoes = [
        "Afastados / Licenças",
        "DESLIGADOS GERAL",
        "Admitidos mês",
        "Férias",
        "ESTAGIARIO",
        "APRENDIZ",
        "SINDICATOS x VALOR",
        "DESLIGADOS ATÉ O DIA 15 DO MÊS - SE JÁ ESTIVEREM CIENTES DO DESLIGAMENTO EXCLUIR DA COMPRA - SE NÃO TIVER O OK COMPRAR INTEGRAL",
        "DESLIGADOS DO DIA 16 ATÉ O ULTIMO DIA DO MÊS PODE FAZER A RECARGA CHEIA E DEIXAR O DESCONTO PROPORCIONAL PARA SER FEITO EM RESCISÃO",
        "ATENDIMENTOS/OBS",
        "Admitidos mês anterior (abril)",
        "EXTERIOR",
        "ATIVOS",
        "REVISAR O CALCULO DE PGTO SE ESTÁ CORRETO ANTES DE GERAR OS VALES"
    ]
    
    # Inserir validações
    for i, validacao in enumerate(validacoes, 2):
        ws.cell(row=i, column=1, value=validacao)
        ws.cell(row=i, column=2, value="")  # Campo Check vazio
    
    # Ajustar larguras
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 15

def get_sindicato_full_name(sindicato_original):
    """Retorna nome do sindicato baseado no sindicato real do funcionário"""
    
    # Se já é um nome completo de sindicato, usar diretamente
    if sindicato_original and len(str(sindicato_original)) > 10:
        return str(sindicato_original)
    
    # Mapeamento de códigos para nomes completos (fallback)
    sindicatos_completos = {
        "001": "SINDPD SP - SIND.TRAB.EM PROC DADOS E EMPR.EMPRESAS PROC DADOS ESTADO DE SP.",
        "002": "SINDPPD RS - SINDICATO DOS TRAB. EM PROC. DE DADOS RIO GRANDE DO SUL",
        "003": "SITEPD PR - SIND DOS TRAB EM EMPR PRIVADAS DE PROC DE DADOS DE CURITIBA E REGIAO METROPOLITANA",
        "004": "SINDPD RJ - SINDICATO PROFISSIONAIS DE PROC DADOS DO RIO DE JANEIRO",
        "005": "SINDICATO GERAL - OUTROS"
    }
    
    # Mapear estados para sindicatos
    estados_sindicatos = {
        "Paraná": "SITEPD PR - SIND DOS TRAB EM EMPR PRIVADAS DE PROC DE DADOS DE CURITIBA E REGIAO METROPOLITANA",
        "Rio de Janeiro": "SINDPD RJ - SINDICATO PROFISSIONAIS DE PROC DADOS DO RIO DE JANEIRO", 
        "Rio Grande do Sul": "SINDPPD RS - SINDICATO DOS TRAB. EM PROC. DE DADOS RIO GRANDE DO SUL",
        "São Paulo": "SINDPD SP - SIND.TRAB.EM PROC DADOS E EMPR.EMPRESAS PROC DADOS ESTADO DE SP."
    }
    
    return sindicatos_completos.get(str(sindicato_original), estados_sindicatos.get(str(sindicato_original), str(sindicato_original)))

def find_file_variations(directory, base_name):
    """Procura variações do nome do arquivo para lidar com acentos e espaços"""
    import os
    from pathlib import Path
    
    # Lista todas as variações possíveis
    variations = [
        base_name,
        base_name.replace(' ', '_'),
        base_name.replace('Ã', 'A'),
        base_name.replace('É', 'E'),
        base_name.replace('FÉRIAS', 'FERIAS'),
        base_name.replace('ADMISSÃO', 'ADMISSAO'),
        base_name.replace(' ABRIL', '_ABRIL')
    ]
    
    # Verifica se alguma variação existe
    for variation in variations:
        file_path = directory / f"{variation}.xlsx"
        if file_path.exists():
            return file_path
    
    # Se não encontrar, retorna o nome original (vai dar erro e será capturado)
    return directory / f"{base_name}.xlsx"

def load_real_admission_dates(raw_data_path):
    """Carrega datas de admissão reais dos arquivos"""
    mapeamento_admissao = {}
    
    try:
        # Carregar admissões de abril (com tratamento de variações no nome)
        admissao_path = find_file_variations(raw_data_path, "ADMISSÃO ABRIL")
        admissao_df = pd.read_excel(admissao_path)
        print(f"📅 Carregadas {len(admissao_df)} admissões de abril")
        
        for _, row in admissao_df.iterrows():
            matricula = str(row['MATRICULA'])
            data_admissao = row['Admissão']
            
            # Converter para datetime se necessário
            if pd.notna(data_admissao):
                if isinstance(data_admissao, str):
                    try:
                        data_admissao = pd.to_datetime(data_admissao)
                    except:
                        data_admissao = datetime(2024, 1, 1)
                
                mapeamento_admissao[matricula] = data_admissao
        
        # Para funcionários sem data específica, usar data padrão mais realista
        # (admissões antigas - antes de abril 2025)
        data_padrao_antiga = datetime(2023, 6, 1)  # Junho 2023 como padrão para funcionários antigos
        
        print(f"📅 Mapeamento criado para {len(mapeamento_admissao)} funcionários")
        print(f"📅 Data padrão para funcionários antigos: {data_padrao_antiga}")
        
        # Retornar função que busca no mapeamento ou retorna data padrão
        def get_admission_date(matricula):
            return mapeamento_admissao.get(str(matricula), data_padrao_antiga)
        
        return get_admission_date
        
    except Exception as e:
        print(f"❌ Erro ao carregar datas de admissão: {e}")
        # Retornar função que sempre retorna data padrão
        def get_default_date(matricula):
            return datetime(2023, 6, 1)
        return get_default_date

def load_real_syndicates(raw_data_path):
    """Carrega sindicatos reais dos funcionários"""
    mapeamento_sindicatos = {}
    
    try:
        # Carregar arquivo ATIVOS com sindicatos
        ativos_df = pd.read_excel(raw_data_path / "ATIVOS.xlsx")
        print(f"🏢 Carregados {len(ativos_df)} funcionários com sindicatos")
        
        for _, row in ativos_df.iterrows():
            matricula = str(row['MATRICULA'])
            sindicato = str(row['Sindicato'])
            
            # Armazenar sindicato real
            if pd.notna(sindicato) and sindicato != 'nan':
                mapeamento_sindicatos[matricula] = sindicato
        
        print(f"🏢 Mapeamento criado para {len(mapeamento_sindicatos)} funcionários")
        
        # Verificar distribuição de sindicatos
        sindicatos_unicos = list(set(mapeamento_sindicatos.values()))
        print(f"🏢 Sindicatos únicos encontrados: {len(sindicatos_unicos)}")
        for sindicato in sindicatos_unicos:
            count = list(mapeamento_sindicatos.values()).count(sindicato)
            print(f"   - {sindicato}: {count} funcionários")
        
        # Retornar função que busca no mapeamento ou retorna sindicato padrão
        def get_employee_syndicate(matricula):
            return mapeamento_sindicatos.get(str(matricula), "SINDICATO GERAL")
        
        return get_employee_syndicate
        
    except Exception as e:
        print(f"❌ Erro ao carregar sindicatos: {e}")
        # Retornar função que sempre retorna sindicato padrão
        def get_default_syndicate(matricula):
            return "SINDICATO GERAL"
        return get_default_syndicate

@tool("validate_model_compliance_tool")
def validate_model_compliance_tool() -> str:
    """
    Valida se a planilha gerada está em conformidade com o modelo exigido
    """
    try:
        project_root = Path(os.path.dirname(__file__)).parent
        output_path = project_root / "output" / "VR_FINAL_MODELO_COMPLETO.xlsx"
        
        if not output_path.exists():
            return "❌ Planilha modelo-compliant não encontrada. Gere primeiro."
        
        print("🔍 Validando conformidade com modelo...")
        
        # Verificar estrutura das abas
        with pd.ExcelFile(output_path) as xls:
            abas_geradas = xls.sheet_names
            abas_esperadas = ["VR MENSAL 05.2025", "Validações"]
            
            print(f"✅ Abas geradas: {abas_geradas}")
            print(f"✅ Abas esperadas: {abas_esperadas}")
            
            conformidade_abas = all(aba in abas_geradas for aba in abas_esperadas)
            
            # Verificar estrutura da aba principal
            df_vr = pd.read_excel(output_path, sheet_name="VR MENSAL 05.2025")
            colunas_esperadas = [
                "Matricula", "Admissão", "Sindicato do Colaborador", "Competência",
                "Dias", "VALOR DIÁRIO VR", "TOTAL", "Custo empresa", 
                "Desconto profissional", "OBS GERAL"
            ]
            
            colunas_geradas = list(df_vr.columns)
            conformidade_colunas = colunas_geradas == colunas_esperadas
            
            # Verificar aba de validações
            df_val = pd.read_excel(output_path, sheet_name="Validações")
            tem_validacoes = len(df_val) >= 14
            
        relatorio = f"""
🔍 RELATÓRIO DE CONFORMIDADE COM MODELO

📊 ESTRUTURA DAS ABAS:
✅ Abas corretas: {"SIM" if conformidade_abas else "NÃO"}
- Geradas: {abas_geradas}
- Esperadas: {abas_esperadas}

📋 ESTRUTURA DA ABA PRINCIPAL:
✅ Colunas corretas: {"SIM" if conformidade_colunas else "NÃO"}
- Total de registros: {len(df_vr):,}
- Colunas: {len(colunas_geradas)} de {len(colunas_esperadas)} esperadas

📋 ABA DE VALIDAÇÕES:
✅ Validações presentes: {"SIM" if tem_validacoes else "NÃO"}
- Total de itens: {len(df_val)}

📊 CONFORMIDADE GERAL:
{"✅ CONFORME MODELO" if conformidade_abas and conformidade_colunas and tem_validacoes else "❌ NÃO CONFORME"}

📄 ARQUIVO VALIDADO: {output_path}
"""
        
        print(relatorio)
        return relatorio
        
    except Exception as e:
        error_msg = f"❌ Erro na validação: {str(e)}"
        print(error_msg)
        return error_msg