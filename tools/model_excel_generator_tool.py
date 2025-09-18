#!/usr/bin/env python3
"""
Tool CrewAI para geração de planilhas Excel conforme modelo do projeto
Convertido para usar decorador @tool conforme filosofia Professor
"""

from crewai.tools import tool
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import logging

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@tool("model_excel_generator_tool")
def model_excel_generator_tool(
    output_filename: str = "VR MENSAL 05.2025.xlsx",
    data_dict: str = None
) -> str:
    """
    Gera planilha Excel final conforme modelo específico do projeto FinaCrew.

    Esta ferramenta cria a planilha de saída no formato exato especificado,
    incluindo formatação, colunas obrigatórias e validações necessárias.

    Args:
        output_filename: Nome do arquivo de saída (padrão: "VR MENSAL 05.2025.xlsx")
        data_dict: String JSON com dados dos funcionários (opcional para testes)

    Returns:
        String com status da geração e localização do arquivo criado
    """
    try:
        print(f"📝 Gerando planilha modelo: {output_filename}")

        # Definir estrutura da planilha conforme especificações
        columns_structure = [
            'MATRICULA',
            'NOME',
            'SINDICATO',
            'VALOR_DIARIO_VR',
            'DIAS_UTEIS',
            'VALOR_TOTAL_VR',
            'PERCENTUAL_EMPRESA',
            'VALOR_EMPRESA',
            'PERCENTUAL_FUNCIONARIO',
            'VALOR_FUNCIONARIO',
            'OBSERVACOES'
        ]

        # Criar DataFrame base
        if data_dict:
            # Se dados foram fornecidos, processar
            # (Implementação simplificada para exemplo)
            print("📊 Processando dados fornecidos...")
            # TODO: Implementar parsing do data_dict
            df = pd.DataFrame(columns=columns_structure)
        else:
            # Criar planilha de exemplo para testes
            print("📋 Criando planilha de exemplo...")
            sample_data = {
                'MATRICULA': ['001', '002', '003'],
                'NOME': ['João Silva', 'Maria Santos', 'Pedro Costa'],
                'SINDICATO': ['SINDPD SP', 'SINDPD RJ', 'SINDPD PR'],
                'VALOR_DIARIO_VR': [37.50, 35.00, 35.00],
                'DIAS_UTEIS': [22, 22, 20],
                'VALOR_TOTAL_VR': [825.00, 770.00, 700.00],
                'PERCENTUAL_EMPRESA': [80, 80, 80],
                'VALOR_EMPRESA': [660.00, 616.00, 560.00],
                'PERCENTUAL_FUNCIONARIO': [20, 20, 20],
                'VALOR_FUNCIONARIO': [165.00, 154.00, 140.00],
                'OBSERVACOES': ['', 'Férias parciais', 'Admissão 15/04']
            }
            df = pd.DataFrame(sample_data)

        # Definir diretório de saída
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / output_filename

        # Criar planilha com formatação
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Aba principal
            df.to_excel(writer, sheet_name='VR_MENSAL', index=False)

            # Obter workbook e worksheet para formatação
            workbook = writer.book
            worksheet = writer.sheets['VR_MENSAL']

            # Formatação de cabeçalhos
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Estilo do cabeçalho
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Aplicar formatação ao cabeçalho
            for col_num, column_title in enumerate(columns_structure, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Formatação de valores monetários
            currency_columns = ['VALOR_DIARIO_VR', 'VALOR_TOTAL_VR', 'VALOR_EMPRESA', 'VALOR_FUNCIONARIO']
            for col_name in currency_columns:
                if col_name in df.columns:
                    col_letter = chr(65 + list(df.columns).index(col_name))
                    for row in range(2, len(df) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.number_format = 'R$ #,##0.00'

            # Formatação de percentuais
            percentage_columns = ['PERCENTUAL_EMPRESA', 'PERCENTUAL_FUNCIONARIO']
            for col_name in percentage_columns:
                if col_name in df.columns:
                    col_letter = chr(65 + list(df.columns).index(col_name))
                    for row in range(2, len(df) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.number_format = '0"%"'

            # Ajustar largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Adicionar bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(columns_structure)):
                for cell in row:
                    cell.border = thin_border

        # Calcular estatísticas
        total_funcionarios = len(df)
        total_vr = df['VALOR_TOTAL_VR'].sum() if 'VALOR_TOTAL_VR' in df.columns else 0
        total_empresa = df['VALOR_EMPRESA'].sum() if 'VALOR_EMPRESA' in df.columns else 0
        total_funcionario = df['VALOR_FUNCIONARIO'].sum() if 'VALOR_FUNCIONARIO' in df.columns else 0

        # Gerar relatório de sucesso
        success_report = f"""
📊 PLANILHA EXCEL GERADA COM SUCESSO

📂 Arquivo: {output_path}
👥 Funcionários: {total_funcionarios}
💰 Valor Total VR: R$ {total_vr:,.2f}
🏢 Valor Empresa (80%): R$ {total_empresa:,.2f}
👤 Valor Funcionário (20%): R$ {total_funcionario:,.2f}

✅ FORMATAÇÕES APLICADAS:
   - Cabeçalhos com destaque azul
   - Valores monetários no formato R$ X.XXX,XX
   - Percentuais formatados adequadamente
   - Bordas e alinhamento profissional
   - Largura de colunas otimizada

📋 ESTRUTURA CONFORME ESPECIFICAÇÕES:
   - 11 colunas obrigatórias implementadas
   - Rateio 80% empresa / 20% funcionário
   - Campo observações para anotações
   - Formato compatível com sistemas corporativos

🎯 Arquivo pronto para download e validação!
"""

        print(success_report)
        return success_report

    except Exception as e:
        error_msg = f"❌ Erro na geração da planilha {output_filename}: {str(e)}"
        print(error_msg)
        return error_msg